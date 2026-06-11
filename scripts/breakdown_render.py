#!/usr/bin/env python3
"""
Phase B of the Mugi director video-shot-breakdown skill.

Takes the Phase-A manifest (per-shot timecode + strip path) plus the breakdown
rows Mugi filled in the middle vision step, and produces a Schema-v8 (audio cut,
10-col) xlsx with the trajectory strip embedded inline on each shot row. Then it
creates a PER-JOB Drive folder under the docgen folder and uploads the xlsx +
every strip + manifest.json into it, so the team can browse the raw contact
sheets and re-render without re-extracting.

It does NOT clean the work dir -- the playbook does that AFTER this step uploads
the folder, because the strips must still exist on disk to be uploaded here.

Pipeline (deterministic, no reasoning):
  1. Load manifest.json (shots + strip paths + title) and rows.json (Mugi's
     filled columns, keyed by shot number).
  2. Build the workbook: header row + one row per shot. Shot # / Timecode come
     from the manifest; the 8 description columns come from rows.json; the strip
     is scaled down and embedded inline.
  3. Save the xlsx into the work dir.
  4. Unless --no-upload: create_folder(<job name>) under docgen, set it to
     "anyone with the link can edit", upload xlsx + all strips + contact sheet +
     manifest.json into it, and drop the individual shot frames into an
     "individual-frames" subfolder (xlsx stays strip-only; frames are for
     on-demand reference). Return the folder + xlsx links + share confirmation.

Usage (CLI):
  python3 scripts/breakdown_render.py --manifest <manifest.json> --rows <rows.json> \
      [--title NAME] [--folder-name NAME] [--out <xlsx path>] \
      [--parent <drive-folder-id>] [--no-upload]

rows.json = a JSON array of objects, one per shot:
  [{"shot": 1, "subject": "...", "live_action": "...", "framing": "...",
    "editing": "...", "vfx": "...", "mg_text": "...", "transition": "...",
    "note": ""}, ...]
"shot" must be the manifest shot number. If EVERY object omits "shot", rows
are matched to shots positionally (array order = manifest order) and the
result carries a warning. Shots missing from rows.json still get a row
(strip + timecode from manifest, description cells blank) -- never silently
dropped.

Output: single line JSON to stdout. On error: {"status":"error","error":...},
exit code 0 (caller parses status). Never prints secret values.
"""

import argparse
import json
import os
import re
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

# Schema v8 minus the two audio columns. The strip is an extra visual column
# (the dreamlin76 thread's embedded-thumbnail idea); it is not one of the 10.
# Display header  ->  rows.json key
COLUMNS = [
    ("Shot #",                     None),          # from manifest
    ("Strip",                      None),          # embedded image
    ("Timecode(s)",                None),          # from manifest
    ("Subject",                    "subject"),
    ("Live Action Description",    "live_action"),
    ("Framing, Angle & Motion",    "framing"),
    ("Editing effects",            "editing"),
    ("VFX & Behavior",             "vfx"),
    ("MG, Text & Animation",       "mg_text"),
    ("Transition out",             "transition"),
    ("Kary's Note",                "note"),
]

# Per-column Excel width (chars). Strip is wide to fit the embedded contact sheet.
COL_WIDTHS = [7, 100, 17, 20, 34, 30, 20, 24, 24, 18, 22]

EMBED_W = 700          # px width the strip is scaled to before embedding
HEADER_FILL = "1F2430"
HEADER_FONT = "FFFFFF"
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_name(s, fallback="breakdown"):
    s = re.sub(r"[^\w\-. ]+", "", (s or "").strip()).strip()
    s = re.sub(r"\s+", " ", s)
    return s[:80] or fallback


def _resolve_strip(strip_path, work_dir):
    """Find the strip image on THIS host.

    The manifest stores container-absolute strip paths. When render runs on a
    different host than extract (or the work dir moved), that absolute path is
    gone -- the old code silently skipped the embed and still returned ok, so
    the xlsx shipped with a blank Strip column. Fall back to the basename under
    work_dir/strips/ before giving up. Returns a usable path or None.
    """
    if strip_path and os.path.exists(strip_path):
        return strip_path
    if strip_path:
        cand = Path(work_dir) / "strips" / os.path.basename(strip_path)
        if cand.exists():
            return str(cand)
    return None


def _scaled_strip(strip_path, tmp_dir):
    """Write a width-EMBED_W copy of the strip for inline embed. Returns (path, w, h)."""
    im = PILImage.open(strip_path)
    if im.mode != "RGB":
        im = im.convert("RGB")
    w, h = im.size
    if w > EMBED_W:
        h = int(h * EMBED_W / w)
        w = EMBED_W
        im = im.resize((w, h), PILImage.LANCZOS)
    out = Path(tmp_dir) / (Path(strip_path).stem + "_embed.jpg")
    im.save(out, quality=85)
    return str(out), w, h


def build_workbook(manifest, rows_by_shot, embed_tmp, work_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Shot Breakdown"

    head_fill = PatternFill("solid", fgColor=HEADER_FILL)
    head_font = Font(bold=True, color=HEADER_FONT, size=11)
    for c, (label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c - 1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    wrap_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    strip_col_letter = get_column_letter(2)

    r = 2
    n_embedded = 0
    for shot in manifest["shots"]:
        sn = shot["shot"]
        row = rows_by_shot.get(sn, {})
        timecode = f"{shot['start_tc']} → {shot['end_tc']}  ({shot['duration_sec']}s)"

        ws.cell(row=r, column=1, value=sn).alignment = center
        ws.cell(row=r, column=3, value=timecode).alignment = wrap_top
        for c, (_, key) in enumerate(COLUMNS, 1):
            if key is None:
                continue
            ws.cell(row=r, column=c, value=row.get(key, "")).alignment = wrap_top

        # Embed the scaled strip inline; size the row + col to it.
        strip_path = _resolve_strip(shot.get("strip"), work_dir)
        row_h = 90
        if strip_path:
            embed_path, ew, eh = _scaled_strip(strip_path, embed_tmp)
            img = XLImage(embed_path)
            img.width, img.height = ew, eh
            img.anchor = f"{strip_col_letter}{r}"
            ws.add_image(img)
            row_h = eh * 0.75 + 8          # px -> points, plus a little padding
            n_embedded += 1
        ws.row_dimensions[r].height = max(row_h, 60)

        for c in range(1, len(COLUMNS) + 1):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    return wb, n_embedded


def render(manifest_path, rows_path, title, folder_name, out_path, parent, do_upload):
    with open(manifest_path) as f:
        manifest = json.load(f)
    rows = []
    rows_warning = None
    if rows_path and os.path.exists(rows_path):
        with open(rows_path) as f:
            rows = json.load(f)
    elif rows_path:
        rows_warning = (f"rows file not found: {rows_path} -- ALL text columns "
                        f"are BLANK.")
    rows_by_shot = {int(r["shot"]): r for r in rows
                    if isinstance(r, dict) and "shot" in r}
    if rows and not rows_by_shot:
        # The vision-fill step writes the array in manifest order but may omit
        # the "shot" key entirely (it happened: every object dropped, xlsx
        # shipped with all text columns blank and status ok). Match
        # positionally against the manifest instead of shipping an empty
        # breakdown -- but say so loudly, alignment is assumed not proven.
        shot_nums = [s["shot"] for s in manifest["shots"]]
        rows_by_shot = {sn: r for sn, r in zip(shot_nums, rows)
                        if isinstance(r, dict)}
        rows_warning = ('rows.json objects have no "shot" key -- rows matched '
                        'to shots POSITIONALLY (array order = manifest order). '
                        'Verify the text columns line up with the strips.')
        if len(rows) != len(shot_nums):
            rows_warning += (f" COUNT MISMATCH: {len(rows)} rows vs "
                             f"{len(shot_nums)} shots; extras dropped.")
    elif rows and len(rows_by_shot) < len(rows):
        rows_warning = (f'only {len(rows_by_shot)} of {len(rows)} rows.json '
                        f'objects had a usable "shot" key; the rest were '
                        f'dropped -- those rows show blank text columns.')

    work_dir = Path(manifest.get("work_dir") or os.path.dirname(manifest_path))
    job_title = title or manifest.get("source", {}).get("title") or "breakdown"
    if not out_path:
        out_path = str(work_dir / f"{_safe_name(job_title)} - breakdown.xlsx")

    embed_tmp = tempfile.mkdtemp(prefix="breakdown_embed_", dir=str(work_dir))
    wb, n_embedded = build_workbook(manifest, rows_by_shot, embed_tmp, work_dir)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)

    n_shots = len(manifest["shots"])
    result = {
        "status": "ok",
        "xlsx_path": os.path.abspath(out_path),
        "n_shots": n_shots,
        "n_rows_filled": len(rows_by_shot),
        "n_strips_embedded": n_embedded,
        "uploaded": False,
    }
    # No-silent-skip: surface every degraded-output condition loudly instead of
    # shipping a "looks fine" ok with empty columns.
    warnings = []
    if rows_warning:
        warnings.append(rows_warning)
    if n_shots and n_embedded == 0:
        warnings.append(
            f"0 of {n_shots} strips embedded -- strip images not found on this host "
            f"(looked in manifest paths + {work_dir}/strips/). Strip column is BLANK."
        )
    if warnings:
        result["warning"] = " | ".join(warnings)

    if do_upload:
        # Lazy import so a --no-upload render needs no Drive creds.
        from breakdown_gdrive import (
            create_folder, upload_file, get_drive_service, set_anyone_writer,
        )
        svc = get_drive_service()
        fname = _safe_name(folder_name or job_title)
        folder = create_folder(fname, parent_id=parent, service=svc)
        fid = folder["id"]
        # "Anyone with the link can edit" on the per-job folder; cascades to every
        # child (xlsx, strips, frames subfolder, contact sheet) in one call.
        share = set_anyone_writer(fid, service=svc)

        up_xlsx = upload_file(out_path, parent_id=fid, service=svc)
        uploaded = [os.path.basename(out_path)]
        # All strips + the manifest, so the team can re-render without re-extracting.
        strips_dir = work_dir / "strips"
        if strips_dir.is_dir():
            for strip in sorted(strips_dir.glob("*.jpg")):
                upload_file(str(strip), parent_id=fid, service=svc)
                uploaded.append(strip.name)
        # Contact sheet (whole-video-at-a-glance) -> job folder, alongside strips.
        contact = manifest.get("contact_sheet")
        if contact and os.path.exists(contact):
            upload_file(contact, parent_id=fid, name="contact_sheet.jpg", service=svc)
            uploaded.append("contact_sheet.jpg")
        # Individual shot frames -> their OWN subfolder. Strip and individual-frame
        # are distinct concepts (a strip groups several actions of one shot); the
        # xlsx stays strip-only and these frames are fetched on demand for reference.
        frames_meta = None
        frames_dir = Path(manifest.get("frames_dir") or (work_dir / "frames"))
        if frames_dir.is_dir():
            # shot_XXX_fNN.jpg = sampled frames; exclude the *_frame.jpg repr copies
            # (those only exist to build the contact sheet, not for browsing).
            frame_imgs = sorted(
                p for p in frames_dir.glob("shot_*_f[0-9]*.jpg")
                if not p.stem.endswith("_frame")
            )
            if frame_imgs:
                sub = create_folder("individual-frames", parent_id=fid, service=svc)
                for fr in frame_imgs:
                    upload_file(str(fr), parent_id=sub["id"], service=svc)
                frames_meta = {
                    "id": sub["id"], "link": sub.get("link"), "n_frames": len(frame_imgs),
                }

        result.update({
            "uploaded": True,
            "folder": {"id": fid, "name": folder.get("name"), "link": folder.get("link")},
            "shared": share,
            "xlsx": {"id": up_xlsx["id"], "link": up_xlsx["link"]},
            "frames_subfolder": frames_meta,
            "uploaded_files": uploaded,
        })

    return result


def _cli():
    ap = argparse.ArgumentParser()
    ap.add_argument("--manifest", required=True)
    ap.add_argument("--rows", default=None)
    ap.add_argument("--title", default=None)
    ap.add_argument("--folder-name", default=None, dest="folder_name")
    ap.add_argument("--out", default=None)
    ap.add_argument("--parent", default=None)
    ap.add_argument("--no-upload", action="store_true", dest="no_upload")
    a = ap.parse_args()
    result = render(a.manifest, a.rows, a.title, a.folder_name, a.out,
                    a.parent, not a.no_upload)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    try:
        _cli()
    except Exception as e:  # noqa: BLE001 - CLI surfaces all errors as JSON
        print(json.dumps({"status": "error", "error": f"{type(e).__name__}: {e}"},
                         ensure_ascii=False))
